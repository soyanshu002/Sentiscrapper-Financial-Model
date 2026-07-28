import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference, Series
from typing import Dict, Any

class ExcelReportGenerator:
    @staticmethod
    def generate_workbook(data: Dict[str, Any]) -> io.BytesIO:
        """
        Generates an executive-grade, multi-tab Excel financial model workbook (.xlsx)
        featuring interactive native Excel charts (Line & Bar charts), corporate KPI cards,
        ML price projections, model backtesting visuals, and multi-channel social sentiment analytics.
        """
        wb = openpyxl.Workbook()

        # -------------------------------------------------------------
        # COLOR PALETTE & DESIGN SYSTEM
        # -------------------------------------------------------------
        NAVY_FILL = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")     # Slate 900
        BLUE_HEADER = PatternFill(start_color="1E3A8A", end_color="1E3A8A", fill_type="solid")   # Blue 900
        SUBHEADER_FILL = PatternFill(start_color="334155", end_color="334155", fill_type="solid")# Slate 700
        KPI_BG_FILL = PatternFill(start_color="F8FAFC", end_color="F8FAFC", fill_type="solid")   # Slate 50
        
        BULLISH_FILL = PatternFill(start_color="DCFCE7", end_color="DCFCE7", fill_type="solid")  # Emerald 100
        BEARISH_FILL = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")  # Rose 100
        NEUTRAL_FILL = PatternFill(start_color="FEF9C3", end_color="FEF9C3", fill_type="solid")  # Amber 100

        TITLE_FONT = Font(name="Calibri", size=15, bold=True, color="FFFFFF")
        SUBTITLE_FONT = Font(name="Calibri", size=10, italic=True, color="CBD5E1")
        HEADER_FONT = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        SECTION_FONT = Font(name="Calibri", size=12, bold=True, color="1E3A8A")
        BOLD_FONT = Font(name="Calibri", size=10, bold=True, color="0F172A")
        REGULAR_FONT = Font(name="Calibri", size=10, color="334155")

        THIN_BORDER = Border(
            left=Side(style='thin', color='CBD5E1'),
            right=Side(style='thin', color='CBD5E1'),
            top=Side(style='thin', color='CBD5E1'),
            bottom=Side(style='thin', color='CBD5E1')
        )

        ticker = data.get("ticker", "UNKNOWN")
        model_type = data.get("model_type", "Random Forest")
        avg_sentiment = data.get("average_sentiment", 0.0)
        rec_text = data.get("recommendation", "")
        metrics = data.get("metrics", {})
        forecast_list = data.get("forecast_data", [])
        hist_list = data.get("historical_data", [])
        sentiment_details = data.get("sentiment_details", [])

        # Parse recommendation signal
        rec_signal = "HOLD"
        for signal in ["BUY", "ACCUMULATE / WEAK BUY", "SELL", "REDUCE / WEAK SELL", "HOLD / NEUTRAL"]:
            if f"`{signal}`" in rec_text or f"advice: {signal}" in rec_text.lower():
                rec_signal = signal
                break

        # -------------------------------------------------------------
        # HELPER SHEET: BACKTEST ANALYTICS DATA (for Charting)
        # -------------------------------------------------------------
        ws_bt = wb.active
        ws_bt.title = "Backtest Data"
        ws_bt.views.sheetView[0].showGridLines = True
        
        ws_bt["A1"] = "Step"
        ws_bt["B1"] = "Actual Test Price"
        ws_bt["C1"] = "Predicted Test Price"

        test_act = metrics.get("test_actual", [])
        test_pred = metrics.get("test_predicted", [])

        for i in range(len(test_act)):
            ws_bt.cell(row=i+2, column=1, value=i+1)
            ws_bt.cell(row=i+2, column=2, value=test_act[i])
            ws_bt.cell(row=i+2, column=3, value=test_pred[i] if i < len(test_pred) else test_act[i])

        # -------------------------------------------------------------
        # TAB 2: ML FORECASTS & METRICS DATA (for Charting & View)
        # -------------------------------------------------------------
        ws_fc = wb.create_sheet(title="ML Forecasts & Metrics")
        ws_fc.views.sheetView[0].showGridLines = True

        ws_fc["A1"] = f"{ticker} - 5-Day Machine Learning Price Projections"
        ws_fc["A1"].font = Font(name="Calibri", size=14, bold=True, color="0F172A")

        ws_fc["A3"] = "MODEL VALIDATION METRICS"
        ws_fc["A3"].font = SECTION_FONT

        met_headers = ["Metric Parameter", "Score", "Description / Benchmark"]
        for col_num, h_text in enumerate(met_headers, 1):
            cell = ws_fc.cell(row=4, column=col_num, value=h_text)
            cell.fill = BLUE_HEADER
            cell.font = HEADER_FONT

        met_data = [
            ("Mean Squared Error (MSE)", metrics.get("mse", 0.0), "Average squared forecast variance"),
            ("Mean Absolute Error (MAE)", metrics.get("mae", 0.0), "Average absolute error magnitude"),
            ("R-Squared (R²)", metrics.get("r2", 0.0), "Percentage of return variance explained (0.0 to 1.0)"),
            ("Directional Accuracy", f"{metrics.get('directional_accuracy', 0.0)*100:.2f}%", "Percentage of correct return direction predictions")
        ]

        for r_i, (m_name, m_val, m_desc) in enumerate(met_data, 5):
            ws_fc.cell(row=r_i, column=1, value=m_name).font = BOLD_FONT
            ws_fc.cell(row=r_i, column=2, value=m_val).font = BOLD_FONT
            ws_fc.cell(row=r_i, column=3, value=m_desc).font = REGULAR_FONT
            for col in range(1, 4):
                ws_fc.cell(row=r_i, column=col).border = THIN_BORDER

        ws_fc["A11"] = "5-DAY PROJECTED CLOSE PRICES"
        ws_fc["A11"].font = SECTION_FONT

        fc_headers = ["Forecast Date", "Predicted Close Price", "Daily Return %", "Cumulative Growth %"]
        for col_num, h_text in enumerate(fc_headers, 1):
            cell = ws_fc.cell(row=12, column=col_num, value=h_text)
            cell.fill = BLUE_HEADER
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        last_close = hist_list[-1]["Close"] if hist_list and "Close" in hist_list[-1] else 100.0
        prev_p = last_close
        for idx, fc_item in enumerate(forecast_list, 13):
            d_val = fc_item.get("Date", "")
            p_val = fc_item.get("Predicted_Close", 0.0)
            daily_ret = ((p_val - prev_p) / prev_p) * 100 if prev_p else 0.0
            cum_ret = ((p_val - last_close) / last_close) * 100 if last_close else 0.0

            ws_fc.cell(row=idx, column=1, value=d_val).font = REGULAR_FONT
            ws_fc.cell(row=idx, column=2, value=round(p_val, 2)).font = BOLD_FONT
            ws_fc.cell(row=idx, column=3, value=round(daily_ret, 2)).font = REGULAR_FONT
            ws_fc.cell(row=idx, column=4, value=round(cum_ret, 2)).font = BOLD_FONT

            for col in range(1, 5):
                c = ws_fc.cell(row=idx, column=col)
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal="center")
            prev_p = p_val

        # -------------------------------------------------------------
        # TAB 3: TECHNICAL INDICATORS & PRICE HISTORY
        # -------------------------------------------------------------
        ws_tech = wb.create_sheet(title="Technicals & Historical Data")
        ws_tech.views.sheetView[0].showGridLines = True

        ws_tech["A1"] = f"{ticker} - Historical Prices & Technical Indicators"
        ws_tech["A1"].font = Font(name="Calibri", size=14, bold=True, color="0F172A")

        tech_headers = ["Date", "Open", "High", "Low", "Close", "RSI (14)", "MACD", "Weighted Sentiment"]
        for col_num, h_text in enumerate(tech_headers, 1):
            cell = ws_tech.cell(row=3, column=col_num, value=h_text)
            cell.fill = BLUE_HEADER
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for idx, row_item in enumerate(hist_list, 4):
            ws_tech.cell(row=idx, column=1, value=row_item.get("Date", "")).font = REGULAR_FONT
            ws_tech.cell(row=idx, column=2, value=round(row_item.get("Open", 0.0), 2)).font = REGULAR_FONT
            ws_tech.cell(row=idx, column=3, value=round(row_item.get("High", 0.0), 2)).font = REGULAR_FONT
            ws_tech.cell(row=idx, column=4, value=round(row_item.get("Low", 0.0), 2)).font = REGULAR_FONT
            ws_tech.cell(row=idx, column=5, value=round(row_item.get("Close", 0.0), 2)).font = BOLD_FONT
            ws_tech.cell(row=idx, column=6, value=round(row_item.get("RSI", 0.0), 2)).font = REGULAR_FONT
            ws_tech.cell(row=idx, column=7, value=round(row_item.get("MACD", 0.0), 2)).font = REGULAR_FONT
            ws_tech.cell(row=idx, column=8, value=round(row_item.get("Weighted_Sentiment", 0.0), 4)).font = REGULAR_FONT

            for col in range(1, 9):
                c = ws_tech.cell(row=idx, column=col)
                c.border = THIN_BORDER
                c.alignment = Alignment(horizontal="center")

        # Historical Close Price Trend Chart on Tech Tab
        chart_hist = LineChart()
        chart_hist.title = f"Historical Close Price Trend ({ticker})"
        chart_hist.style = 13
        chart_hist.y_axis.title = "Price"
        chart_hist.x_axis.title = "Date"
        chart_hist.height = 12
        chart_hist.width = 18

        h_data_ref = Reference(ws_tech, min_col=5, min_row=3, max_row=3 + len(hist_list))
        h_cats_ref = Reference(ws_tech, min_col=1, min_row=4, max_row=3 + len(hist_list))
        chart_hist.add_data(h_data_ref, titles_from_data=True)
        chart_hist.set_categories(h_cats_ref)
        ws_tech.add_chart(chart_hist, "J3")

        # -------------------------------------------------------------
        # TAB 4: SOCIAL SENTIMENT CORPUS DATA
        # -------------------------------------------------------------
        ws_sent = wb.create_sheet(title="Social Sentiment Corpus")
        ws_sent.views.sheetView[0].showGridLines = True

        ws_sent["A1"] = f"{ticker} - Harvested Social & News Feed Sentiment"
        ws_sent["A1"].font = Font(name="Calibri", size=14, bold=True, color="0F172A")

        sent_headers = ["Platform Source", "Post / Headline Text", "Polarity Compound Score", "Sentiment Classification"]
        for col_num, h_text in enumerate(sent_headers, 1):
            cell = ws_sent.cell(row=3, column=col_num, value=h_text)
            cell.fill = BLUE_HEADER
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

        for idx, s_item in enumerate(sentiment_details, 4):
            src = s_item.get("source", "News")
            txt = s_item.get("text", "")
            sc = s_item.get("compound", 0.0)

            if sc >= 0.05:
                rating = "Bullish"
                rating_fill = BULLISH_FILL
            elif sc <= -0.05:
                rating = "Bearish"
                rating_fill = BEARISH_FILL
            else:
                rating = "Neutral"
                rating_fill = NEUTRAL_FILL

            ws_sent.cell(row=idx, column=1, value=src).font = BOLD_FONT
            ws_sent.cell(row=idx, column=2, value=txt).font = REGULAR_FONT
            ws_sent.cell(row=idx, column=3, value=round(sc, 4)).font = BOLD_FONT
            
            c4 = ws_sent.cell(row=idx, column=4, value=rating)
            c4.font = BOLD_FONT
            c4.fill = rating_fill

            for col in range(1, 5):
                c = ws_sent.cell(row=idx, column=col)
                c.border = THIN_BORDER
                if col in [1, 3, 4]:
                    c.alignment = Alignment(horizontal="center")

        # Sentiment breakdown summary for chart
        ws_sent["F3"] = "Platform"
        ws_sent["G3"] = "Average Compound Polarity"

        platforms = ["Reddit", "Twitter", "Telegram", "News"]
        for r_idx, p in enumerate(platforms, 4):
            items = [s for s in sentiment_details if p.lower() in s.get("source", "").lower()]
            avg_score = sum(s.get("compound", 0.0) for s in items) / len(items) if items else avg_sentiment
            ws_sent.cell(row=r_idx, column=6, value=p)
            ws_sent.cell(row=r_idx, column=7, value=round(avg_score, 4))

        # -------------------------------------------------------------
        # TAB 1: EXECUTIVE DASHBOARD (PRIMARY MAIN SHEET)
        # -------------------------------------------------------------
        ws_exec = wb.create_sheet(title="Executive Dashboard", index=0)
        ws_exec.views.sheetView[0].showGridLines = True

        # Banner Block
        ws_exec.merge_cells("A1:D1")
        ws_exec["A1"] = f"  SentiScrapper AI Financial Valuation & Executive Dashboard"
        ws_exec["A1"].font = TITLE_FONT
        ws_exec["A1"].fill = NAVY_FILL
        ws_exec["A1"].alignment = Alignment(vertical="center")
        ws_exec.row_dimensions[1].height = 32

        ws_exec.merge_cells("A2:D2")
        ws_exec["A2"] = f"  Ticker Symbol: {ticker}  |  ML Engine: {model_type}  |  Social Sentiment Polarity: {avg_sentiment:+.4f}"
        ws_exec["A2"].font = SUBTITLE_FONT
        ws_exec["A2"].fill = NAVY_FILL
        ws_exec["A2"].alignment = Alignment(vertical="center")
        ws_exec.row_dimensions[2].height = 22

        # Key Performance Indicators Table
        ws_exec["A4"] = "EXECUTIVE VALUATION KEY PERFORMANCE INDICATORS"
        ws_exec["A4"].font = SECTION_FONT

        kpi_headers = ["Metric Parameter", "Value / Output", "Benchmark / Context"]
        for col_num, h_text in enumerate(kpi_headers, 1):
            cell = ws_exec.cell(row=5, column=col_num, value=h_text)
            cell.fill = BLUE_HEADER
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")

        kpi_rows = [
            ("Target Stock Ticker", ticker, "NSE / BSE / US Equities"),
            ("Machine Learning Engine", model_type, "Quant Analyst Pipeline"),
            ("Final Advisory Signal", rec_signal, "Portfolio Manager Synthesis"),
            ("Average Social Sentiment Polarity", f"{avg_sentiment:+.4f}", "-1.0 (Bearish) to +1.0 (Bullish)"),
            ("Model Directional Accuracy", f"{metrics.get('directional_accuracy', 0.0)*100:.2f}%", "Backtest Test Window"),
            ("Model Mean Squared Error (MSE)", f"{metrics.get('mse', 0.0):.4f}", "Squared Return Variance"),
            ("Model R² Coefficient", f"{metrics.get('r2', 0.0):.4f}", "Variance Explained (0.0 to 1.0)")
        ]

        for row_idx, (param, val, bench) in enumerate(kpi_rows, 6):
            c1 = ws_exec.cell(row=row_idx, column=1, value=param)
            c2 = ws_exec.cell(row=row_idx, column=2, value=val)
            c3 = ws_exec.cell(row=row_idx, column=3, value=bench)

            c1.font, c2.font, c3.font = REGULAR_FONT, BOLD_FONT, REGULAR_FONT
            c1.border, c2.border, c3.border = THIN_BORDER, THIN_BORDER, THIN_BORDER

            if param == "Final Advisory Signal":
                c2.fill = BULLISH_FILL if "BUY" in val else BEARISH_FILL if "SELL" in val else NEUTRAL_FILL

        # Advisory Synthesis Section
        ws_exec["A15"] = "PORTFOLIO MANAGER ADVISORY REPORT"
        ws_exec["A15"].font = SECTION_FONT

        rec_lines = [line for line in rec_text.split('\n') if line.strip()]
        start_r = 16
        for line in rec_lines:
            ws_exec.cell(row=start_r, column=1, value=line).font = REGULAR_FONT
            start_r += 1

        # -------------------------------------------------------------
        # EMBEDDED NATIVE EXCEL CHARTS ON EXECUTIVE DASHBOARD
        # -------------------------------------------------------------

        # CHART 1: 5-Day ML Projected Price Line Chart
        chart_forecast = LineChart()
        chart_forecast.title = f"5-Day ML Projected Close Price Trajectory ({ticker})"
        chart_forecast.style = 13
        chart_forecast.y_axis.title = "Predicted Price"
        chart_forecast.x_axis.title = "Forecast Date"
        chart_forecast.height = 11
        chart_forecast.width = 17

        fc_data_ref = Reference(ws_fc, min_col=2, min_row=12, max_row=12 + len(forecast_list))
        fc_cats_ref = Reference(ws_fc, min_col=1, min_row=13, max_row=12 + len(forecast_list))
        chart_forecast.add_data(fc_data_ref, titles_from_data=True)
        chart_forecast.set_categories(fc_cats_ref)
        ws_exec.add_chart(chart_forecast, "F4")

        # CHART 2: Quant Model Backtest Actual vs Predicted Line Chart
        chart_backtest = LineChart()
        chart_backtest.title = "Quant Model Backtest: Actual vs. Predicted Prices"
        chart_backtest.style = 10
        chart_backtest.y_axis.title = "Price"
        chart_backtest.x_axis.title = "Test Points"
        chart_backtest.height = 11
        chart_backtest.width = 17

        bt_data_ref = Reference(ws_bt, min_col=2, max_col=3, min_row=1, max_row=1 + len(test_act))
        bt_cats_ref = Reference(ws_bt, min_col=1, min_row=2, max_row=1 + len(test_act))
        chart_backtest.add_data(bt_data_ref, titles_from_data=True)
        chart_backtest.set_categories(bt_cats_ref)
        ws_exec.add_chart(chart_backtest, "F20")

        # CHART 3: Social Platform Polarity Column Bar Chart
        chart_sentiment = BarChart()
        chart_sentiment.type = "col"
        chart_sentiment.style = 11
        chart_sentiment.title = "Social Mining Polarity Score by Channel"
        chart_sentiment.y_axis.title = "Compound Polarity (-1.0 to +1.0)"
        chart_sentiment.x_axis.title = "Platform Channel"
        chart_sentiment.height = 10
        chart_sentiment.width = 16

        st_data_ref = Reference(ws_sent, min_col=7, min_row=3, max_row=3 + len(platforms))
        st_cats_ref = Reference(ws_sent, min_col=6, min_row=4, max_row=3 + len(platforms))
        chart_sentiment.add_data(st_data_ref, titles_from_data=True)
        chart_sentiment.set_categories(st_cats_ref)
        ws_exec.add_chart(chart_sentiment, "A30")

        # Auto-adjust column widths across all worksheets
        for sheet in wb.worksheets:
            for col in sheet.columns:
                max_len = max(len(str(cell.value or '')) for cell in col)
                col_letter = get_column_letter(col[0].column)
                sheet.column_dimensions[col_letter].width = min(max(max_len + 3, 14), 85)

        # Save to memory buffer
        output = io.BytesIO()
        wb.save(output)
        output.seek(0)
        return output
